"""
Hisobot eksporterlari — Excel (openpyxl), PDF (reportlab), grafik (matplotlib).

Og'ir kutubxonalar funksiya ichida lazy import qilinadi — agar kutubxona
o'rnatilmagan bo'lsa, funksiya None qaytaradi (bot baribir matnli hisobot beradi).
Barcha funksiyalar SINXRON — handlerда asyncio.to_thread orqali chaqiriladi.
"""

from __future__ import annotations

import io
import logging
from datetime import datetime
from typing import Any, Callable, Dict, List, Optional

logger = logging.getLogger(__name__)

# Kasallik id → o'qiladigan nom (uz)
try:
    from medical import ALL_DISEASES
    _DISEASE_NAME = {d.id: d.name_uz for d in ALL_DISEASES}
except Exception:  # noqa: BLE001
    _DISEASE_NAME = {}

try:
    from domain.symptoms import question_text as _question_text
except Exception:  # noqa: BLE001
    def _question_text(qid: str, lang: str = "uz") -> str:
        return qid


def disease_name(disease_id: Optional[str]) -> str:
    if not disease_id:
        return "—"
    return _DISEASE_NAME.get(disease_id, disease_id)


def _fmt_dt(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    return str(value or "")


def _username(row: Dict[str, Any]) -> str:
    u = row.get("username")
    return f"@{u}" if u else "—"


# ─────────────────────────────────────────────────────────────────
#  GRAFIK (PNG)
# ─────────────────────────────────────────────────────────────────

def build_overview_chart(series: List[tuple], top_diseases: List[tuple], title: str) -> Optional[bytes]:
    """Kunlik faollik + top kasalliklar grafigi (PNG bytes)."""
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except Exception as exc:  # noqa: BLE001
        logger.warning("matplotlib yo'q: %s", exc)
        return None

    try:
        fig, axes = plt.subplots(1, 2, figsize=(12, 4.5))
        fig.suptitle(title)

        # Chap: kunlik sessiyalar
        ax1 = axes[0]
        if series:
            dates = [d for d, _ in series]
            counts = [c for _, c in series]
            ax1.bar(range(len(dates)), counts, color="#2a9d8f")
            ax1.set_xticks(range(len(dates)))
            ax1.set_xticklabels(dates, rotation=45, ha="right", fontsize=7)
            ax1.set_title("Kunlik sessiyalar")
            ax1.set_ylabel("Soni")
        else:
            ax1.text(0.5, 0.5, "Ma'lumot yo'q", ha="center", va="center")
            ax1.set_title("Kunlik sessiyalar")

        # O'ng: top kasalliklar
        ax2 = axes[1]
        if top_diseases:
            names = [disease_name(d)[:22] for d, _ in top_diseases[:8]][::-1]
            vals = [c for _, c in top_diseases[:8]][::-1]
            ax2.barh(range(len(names)), vals, color="#e76f51")
            ax2.set_yticks(range(len(names)))
            ax2.set_yticklabels(names, fontsize=7)
            ax2.set_title("Eng ko'p tashxislar")
            ax2.set_xlabel("Soni")
        else:
            ax2.text(0.5, 0.5, "Ma'lumot yo'q", ha="center", va="center")
            ax2.set_title("Eng ko'p tashxislar")

        fig.tight_layout(rect=[0, 0, 1, 0.95])
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=120)
        plt.close(fig)
        buf.seek(0)
        return buf.getvalue()
    except Exception as exc:  # noqa: BLE001
        logger.error("build_overview_chart xatoligi: %s", exc)
        return None


# ─────────────────────────────────────────────────────────────────
#  EXCEL (.xlsx)
# ─────────────────────────────────────────────────────────────────

def build_excel_report(
    title: str,
    ov: Dict[str, Any],
    sessions: List[Dict[str, Any]],
    answers: Dict[int, List[Dict[str, Any]]],
    users: List[Dict[str, Any]],
) -> Optional[bytes]:
    """To'liq audit Excel: Summary / Sessions / Answers / Users varaqlari."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except Exception as exc:  # noqa: BLE001
        logger.warning("openpyxl yo'q: %s", exc)
        return None

    try:
        wb = Workbook()

        # 1) Summary
        ws = wb.active
        ws.title = "Summary"
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14)
        rows = [
            ("Jami obunachilar", ov.get("total_users", 0)),
            ("Yangi obunachilar (davr)", ov.get("new_users", 0)),
            ("Sessiyalar (davr)", ov.get("total_sessions", 0)),
            ("Faol foydalanuvchilar (davr)", ov.get("active_users", 0)),
            ("O'rtacha savol soni", round(ov.get("avg_questions", 0.0), 1)),
            ("O'rtacha ishonch", round(ov.get("avg_confidence", 0.0), 2)),
            ("Aniqlanmagan (uncertain)", ov.get("uncertain", 0)),
        ]
        r = 3
        for label, val in rows:
            ws.cell(row=r, column=1, value=label)
            ws.cell(row=r, column=2, value=val)
            r += 1
        r += 1
        ws.cell(row=r, column=1, value="Eng ko'p tashxislar").font = Font(bold=True)
        r += 1
        for did, cnt in ov.get("top_diseases", []):
            ws.cell(row=r, column=1, value=disease_name(did))
            ws.cell(row=r, column=2, value=cnt)
            r += 1

        # 2) Sessions
        ws2 = wb.create_sheet("Sessions")
        headers = ["session_id", "sana", "user_id", "ism", "username", "til",
                   "kategoriya", "yosh", "jins", "joylashuv", "savollar",
                   "tashxis", "ishonch", "muqobil_2", "muqobil_3", "aniqlanmagan"]
        ws2.append(headers)
        for c in ws2[1]:
            c.font = Font(bold=True)
        for s in sessions:
            ws2.append([
                s.get("session_id"), _fmt_dt(s.get("started_at")), s.get("telegram_id"),
                s.get("name"), _username(s), s.get("lang"), s.get("category"),
                s.get("age_group"), s.get("sex"), str(s.get("location") or ""),
                s.get("question_count"), disease_name(s.get("top1_disease")),
                round(float(s.get("top1_conf") or 0), 2),
                disease_name(s.get("top2_disease")) if s.get("top2_disease") else "",
                disease_name(s.get("top3_disease")) if s.get("top3_disease") else "",
                "ha" if s.get("uncertain") else "yo'q",
            ])

        # 3) Answers (to'liq audit)
        ws3 = wb.create_sheet("Answers")
        ws3.append(["session_id", "qadam", "savol_id", "savol", "javob"])
        for c in ws3[1]:
            c.font = Font(bold=True)
        for sid, alist in answers.items():
            for a in alist:
                ws3.append([
                    sid, a.get("step"), a.get("question_id"),
                    _question_text(a.get("question_id", ""), "uz"),
                    "Ha" if a.get("answer") else "Yo'q",
                ])

        # 4) Users
        ws4 = wb.create_sheet("Users")
        ws4.append(["user_id", "ism", "username", "til", "sessiyalar", "ro'yxatdan", "oxirgi_faollik"])
        for c in ws4[1]:
            c.font = Font(bold=True)
        for u in users:
            ws4.append([
                u.get("telegram_id"), u.get("name"), _username(u), u.get("lang"),
                u.get("sessions"), _fmt_dt(u.get("created_at")), _fmt_dt(u.get("last_seen")),
            ])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()
    except Exception as exc:  # noqa: BLE001
        logger.error("build_excel_report xatoligi: %s", exc)
        return None


def build_user_excel(
    user: Dict[str, Any],
    sessions: List[Dict[str, Any]],
    answers: Dict[int, List[Dict[str, Any]]],
) -> Optional[bytes]:
    """Bitta obunachi bo'yicha individual Excel (sessiyalar + savol-javoblar)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except Exception as exc:  # noqa: BLE001
        logger.warning("openpyxl yo'q: %s", exc)
        return None
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Profil"
        ws["A1"] = f"Obunachi hisoboti — {user.get('name','—')}"
        ws["A1"].font = Font(bold=True, size=14)
        info = [
            ("User ID", user.get("telegram_id")),
            ("Ism", user.get("name")),
            ("Username", _username(user)),
            ("Til", user.get("lang")),
            ("Ro'yxatdan", _fmt_dt(user.get("created_at"))),
            ("Oxirgi faollik", _fmt_dt(user.get("last_seen"))),
            ("Sessiyalar soni", len(sessions)),
        ]
        r = 3
        for label, val in info:
            ws.cell(row=r, column=1, value=label)
            ws.cell(row=r, column=2, value=val)
            r += 1

        ws2 = wb.create_sheet("Sessiyalar")
        ws2.append(["session_id", "sana", "kategoriya", "yosh", "jins", "joylashuv",
                    "savollar", "tashxis", "ishonch"])
        for c in ws2[1]:
            c.font = Font(bold=True)
        for s in sessions:
            ws2.append([
                s.get("session_id"), _fmt_dt(s.get("started_at")), s.get("category"),
                s.get("age_group"), s.get("sex"), str(s.get("location") or ""),
                s.get("question_count"), disease_name(s.get("top1_disease")),
                round(float(s.get("top1_conf") or 0), 2),
            ])

        ws3 = wb.create_sheet("Savol-javoblar")
        ws3.append(["session_id", "qadam", "savol", "javob"])
        for c in ws3[1]:
            c.font = Font(bold=True)
        for sid, alist in answers.items():
            for a in alist:
                ws3.append([
                    sid, a.get("step"),
                    _question_text(a.get("question_id", ""), "uz"),
                    "Ha" if a.get("answer") else "Yo'q",
                ])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()
    except Exception as exc:  # noqa: BLE001
        logger.error("build_user_excel xatoligi: %s", exc)
        return None


# ─────────────────────────────────────────────────────────────────
#  PDF
# ─────────────────────────────────────────────────────────────────

def build_pdf_report(
    title: str,
    ov: Dict[str, Any],
    chart_png: Optional[bytes] = None,
) -> Optional[bytes]:
    """Umumiy hisobot PDF (sammari + top tashxislar + grafik)."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image,
        )
        from reportlab.lib.styles import getSampleStyleSheet
    except Exception as exc:  # noqa: BLE001
        logger.warning("reportlab yo'q: %s", exc)
        return None
    try:
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, title=title)
        styles = getSampleStyleSheet()
        story = [Paragraph(title, styles["Title"]), Spacer(1, 0.4 * cm)]

        summary = [
            ["Ko'rsatkich", "Qiymat"],
            ["Jami obunachilar", str(ov.get("total_users", 0))],
            ["Yangi obunachilar (davr)", str(ov.get("new_users", 0))],
            ["Sessiyalar (davr)", str(ov.get("total_sessions", 0))],
            ["Faol foydalanuvchilar (davr)", str(ov.get("active_users", 0))],
            ["O'rtacha savol soni", str(round(ov.get("avg_questions", 0.0), 1))],
            ["O'rtacha ishonch", str(round(ov.get("avg_confidence", 0.0), 2))],
            ["Aniqlanmagan (uncertain)", str(ov.get("uncertain", 0))],
        ]
        tbl = Table(summary, colWidths=[8 * cm, 6 * cm])
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2a9d8f")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
        ]))
        story += [tbl, Spacer(1, 0.5 * cm)]

        if ov.get("top_diseases"):
            story.append(Paragraph("Eng ko'p tashxislar", styles["Heading2"]))
            data = [["Tashxis", "Soni"]] + [
                [disease_name(d), str(c)] for d, c in ov["top_diseases"][:10]
            ]
            t2 = Table(data, colWidths=[10 * cm, 4 * cm])
            t2.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#264653")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
            ]))
            story += [t2, Spacer(1, 0.5 * cm)]

        if chart_png:
            try:
                story.append(Image(io.BytesIO(chart_png), width=17 * cm, height=6.4 * cm))
            except Exception:  # noqa: BLE001
                pass

        doc.build(story)
        buf.seek(0)
        return buf.getvalue()
    except Exception as exc:  # noqa: BLE001
        logger.error("build_pdf_report xatoligi: %s", exc)
        return None


def build_user_pdf(
    user: Dict[str, Any],
    sessions: List[Dict[str, Any]],
    answers: Dict[int, List[Dict[str, Any]]],
) -> Optional[bytes]:
    """Bitta obunachi bo'yicha individual PDF."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        )
        from reportlab.lib.styles import getSampleStyleSheet
    except Exception as exc:  # noqa: BLE001
        logger.warning("reportlab yo'q: %s", exc)
        return None
    try:
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4)
        styles = getSampleStyleSheet()
        story = [Paragraph(f"Obunachi hisoboti — {user.get('name','—')}", styles["Title"])]
        story.append(Paragraph(
            f"ID: {user.get('telegram_id')} | {_username(user)} | "
            f"Til: {user.get('lang','—')} | Sessiyalar: {len(sessions)}",
            styles["Normal"],
        ))
        story.append(Spacer(1, 0.4 * cm))

        for s in sessions:
            head = (
                f"#{s.get('session_id')} — {_fmt_dt(s.get('started_at'))} | "
                f"{s.get('category','—')} | Tashxis: {disease_name(s.get('top1_disease'))} "
                f"({round(float(s.get('top1_conf') or 0), 2)})"
            )
            story.append(Paragraph(head, styles["Heading3"]))
            alist = answers.get(s.get("session_id"), [])
            if alist:
                data = [["Savol", "Javob"]] + [
                    [_question_text(a.get("question_id", ""), "uz")[:70],
                     "Ha" if a.get("answer") else "Yo'q"]
                    for a in alist
                ]
                t = Table(data, colWidths=[13 * cm, 3 * cm])
                t.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2a9d8f")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                ]))
                story += [t, Spacer(1, 0.3 * cm)]

        if not sessions:
            story.append(Paragraph("Bu obunachida sessiyalar yo'q.", styles["Normal"]))

        doc.build(story)
        buf.seek(0)
        return buf.getvalue()
    except Exception as exc:  # noqa: BLE001
        logger.error("build_user_pdf xatoligi: %s", exc)
        return None
